"""
06_build_excel_template.py

Builds a physical inventory / cycle count template -- the genuine manual
input layer in a real SAP MM environment. Store/warehouse staff perform
periodic cycle counts (SAP transactions MI01/MI04/MI07) to reconcile actual
shelf stock against the system's recorded stock; discrepancies get posted
as inventory adjustments. This is NOT a rebuilt dashboard -- it's the
paper/spreadsheet layer that exists before anything hits the system cleanly,
which is exactly why this kind of file has the same real-world messiness
problem the cleaning script in this project solves.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FORMULA_FONT = Font(name=FONT_NAME, color="000000")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------------------
# SHEET 1: Instructions
# ---------------------------------------------------------------------------
ws_intro = wb.active
ws_intro.title = "Instructions"
ws_intro.sheet_view.showGridLines = False
ws_intro.column_dimensions["A"].width = 100

lines = [
    ("Physical Inventory / Cycle Count Sheet", 16, True),
    ("", 11, False),
    ("What this is for", 13, True),
    ("Store and warehouse staff use this sheet to record what's actually on the", 11, False),
    ("shelf during a periodic stock count, and compare it to what the system", 11, False),
    ("(SAP) thinks is there. Differences ('variances') get investigated and then", 11, False),
    ("posted as an inventory adjustment.", 11, False),
    ("", 11, False),
    ("How to use it", 13, True),
    ("1. Go to the 'Count Sheet' tab.", 11, False),
    ("2. For each material_id / plant combination being counted, fill in the", 11, False),
    ("   yellow 'Physical Count' column with what you actually counted.", 11, False),
    ("3. The 'System Stock' column is pulled from the last known system", 11, False),
    ("   balance (fill in manually here, or paste from an export) -- it is NOT", 11, False),
    ("   something store staff should edit.", 11, False),
    ("4. The 'Variance' and 'Variance %' columns calculate automatically.", 11, False),
    ("5. Anything over the tolerance threshold (+/- 5%) is flagged in red for", 11, False),
    ("   investigation before an adjustment is posted.", 11, False),
    ("", 11, False),
    ("Why this matters for the analysis", 13, True),
    ("Cycle counts are how real stock discrepancies get caught -- without them,", 11, False),
    ("a system can report stock that doesn't physically exist (or vice versa),", 11, False),
    ("which would quietly distort every inventory metric in this project.", 11, False),
]
row = 1
for text, size, bold in lines:
    cell = ws_intro.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color="1F3A5F" if bold else "000000")
    row += 1

# ---------------------------------------------------------------------------
# SHEET 2: Lookups
# ---------------------------------------------------------------------------
ws_lookup = wb.create_sheet("Lookups")
PLANTS = ["FR10", "FR90", "UK10", "DE10", "US10", "US20", "AE10", "MA10"]
REASON_CODES = ["Miscount (recount confirmed)", "Damaged / written off", "Theft / shrinkage suspected",
                "Not yet posted in system", "Found misplaced stock", "Other (see notes)"]
ws_lookup["A1"] = "Plant"
ws_lookup["B1"] = "ReasonCode"
for i, v in enumerate(PLANTS, start=2):
    ws_lookup.cell(row=i, column=1, value=v)
for i, v in enumerate(REASON_CODES, start=2):
    ws_lookup.cell(row=i, column=2, value=v)
ws_lookup.sheet_state = "hidden"

# ---------------------------------------------------------------------------
# SHEET 3: Count Sheet
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Count Sheet")
headers = [
    "count_date", "plant", "material_id", "material_description",
    "system_stock", "physical_count", "variance", "variance_pct",
    "reason_code", "notes",
]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = 18
ws.row_dimensions[1].height = 30

example_rows = [
    ["2025-07-01", "FR10", "P416725", "The Revitalizing Hydrating Serum", 42, 40, None, None, "Miscount (recount confirmed)", ""],
    ["2025-07-01", "FR10", "P76000", "Pure Poison", 18, 18, None, None, "", ""],
    ["2025-07-01", "US10", "P162554", "Almond Smoothing Oil", 25, 19, None, None, "Theft / shrinkage suspected", "Flagged to loss prevention"],
]
for r_offset, row_vals in enumerate(example_rows):
    r = r_offset + 2
    for col, v in enumerate(row_vals, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.border = BORDER
        if col in (2, 3, 4, 5, 6, 9, 10):
            c.fill = INPUT_FILL
            c.font = Font(name=FONT_NAME, italic=True, color="666666")
    # variance formula = physical_count - system_stock
    ws.cell(row=r, column=7, value=f"=F{r}-E{r}")
    # variance % = variance / system_stock
    ws.cell(row=r, column=8, value=f"=IF(E{r}=0,\"\",G{r}/E{r})")
    ws.cell(row=r, column=8).number_format = "0.0%"

# Extend formulas + formatting down to row 200
for r in range(2, 201):
    ws.cell(row=r, column=7, value=f"=F{r}-E{r}")
    ws.cell(row=r, column=8, value=f"=IF(E{r}=0,\"\",G{r}/E{r})")
    ws.cell(row=r, column=8).number_format = "0.0%"
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).border = BORDER

# Conditional formatting: flag |variance %| > 5%
from openpyxl.formatting.rule import FormulaRule
red_fill = PatternFill(start_color="F8CBCB", end_color="F8CBCB", fill_type="solid")
ws.conditional_formatting.add(
    "H2:H200",
    FormulaRule(formula=["AND(ISNUMBER(H2),ABS(H2)>0.05)"], fill=red_fill)
)

dv_plant = DataValidation(type="list", formula1="=Lookups!$A$2:$A$9", allow_blank=True)
dv_reason = DataValidation(type="list", formula1="=Lookups!$B$2:$B$7", allow_blank=True)
ws.add_data_validation(dv_plant)
ws.add_data_validation(dv_reason)
dv_plant.add("B2:B200")
dv_reason.add("I2:I200")

ws.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# SHEET 4: Summary (formulas only, references Count Sheet)
# ---------------------------------------------------------------------------
ws_sum = wb.create_sheet("Summary")
ws_sum.column_dimensions["A"].width = 35
ws_sum.column_dimensions["B"].width = 18

summary_rows = [
    ("Total lines counted", "='Count Sheet'!COUNTA(C2:C200)".replace("COUNTA(", "").replace(")", "")),
]
# Build safe formulas directly (avoid the placeholder above)
ws_sum["A1"] = "Cycle Count Summary"
ws_sum["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color="1F3A5F")

labels_formulas = [
    ("Total lines counted", "=COUNTA('Count Sheet'!C2:C200)"),
    ("Lines with a variance flagged (>5%)", "=COUNTIF('Count Sheet'!H2:H200,\">0.05\")+COUNTIF('Count Sheet'!H2:H200,\"<-0.05\")"),
    ("Total system stock (counted lines)", "=SUM('Count Sheet'!E2:E200)"),
    ("Total physical count", "=SUM('Count Sheet'!F2:F200)"),
    ("Net variance (units)", "=SUM('Count Sheet'!G2:G200)"),
]
r = 3
for label, formula in labels_formulas:
    ws_sum.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, size=11)
    fcell = ws_sum.cell(row=r, column=2, value=formula)
    fcell.font = Font(name=FONT_NAME, size=11, bold=True)
    fcell.border = BORDER
    r += 1

wb.save("excel/cycle_count_template.xlsx")
print("Saved excel/cycle_count_template.xlsx")
